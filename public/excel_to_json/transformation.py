import openpyxl
import json

excel_file = './locationNumber.xlsx'
json_file = './data.json'

wb = openpyxl.load_workbook(excel_file, read_only=True)

sheet = wb.worksheets[1]  # second sheet

key_list = []

for col_num in range(1, sheet.max_column + 1):
    key_list.append(sheet.cell(row=1, column=col_num).value)

# print(key_list)

# from excel data to object data

data_list = []  # {}?
for row_num in range(2, sheet.max_row + 1):
    tmp_dict = {}
    for col_num in range(1, sheet.max_column + 1):
        val = sheet.cell(row=row_num, column=col_num).value
        tmp_dict[key_list[col_num - 1]] = val

    # print(tmp_dict)
    data_list.append(tmp_dict)

# print(data_list)

wb.close()

# from object data to json data
with open(json_file, 'w', encoding='utf-8') as fp:
    json.dump(data_list, fp, indent=4, ensure_ascii=False)
